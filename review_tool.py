from pathlib import Path
import json
from datetime import datetime
import shutil

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Не установлен openpyxl.")
    print("Выполни команду: python -m pip install openpyxl")
    raise


ROOT = Path(__file__).resolve().parent

FEEDBACK_DIR = ROOT / "feedback"
REVIEW_DIR = ROOT / "review"
EXPECTED_DIR = ROOT / "expected"
RULES_DIR = ROOT / "rules"
CORPUS_SYNTHETIC_DIR = ROOT / "corpus" / "synthetic"
BACKUP_DIR = ROOT / "archive" / "review_import_backups"

CASES_PATH = FEEDBACK_DIR / "cases.jsonl"
DECISIONS_PATH = FEEDBACK_DIR / "decisions.jsonl"
CANDIDATE_RULES_PATH = FEEDBACK_DIR / "candidate_rules.jsonl"

# ВАЖНО:
# expected/regression_cases.jsonl — стабильный synthetic baseline.
# review_tool.py import больше НЕ перезаписывает этот файл.
BASELINE_REGRESSION_CASES_PATH = EXPECTED_DIR / "regression_cases.jsonl"

# Сюда пишутся тесты, созданные из Excel-review.
REVIEW_GENERATED_CASES_PATH = EXPECTED_DIR / "review_generated_cases.jsonl"

REVIEW_PATH = REVIEW_DIR / "review_cases.xlsx"

MANUAL_HIDE_PATH = RULES_DIR / "manual_hide.txt"
MANUAL_ALLOW_PATH = RULES_DIR / "manual_allow.txt"


VISIBLE_COLUMNS = [
    "№",
    "Кейс",
    "Файл",
    "Блок",
    "Фрагмент для проверки",
    "Что предлагает система",
    "Почему попало в проверку",
    "Решение человека",
    "Область действия",
    "Тип результата",
    "Создать правило-кандидат?",
    "Создать review-test?",
    "Комментарий человека",
]

TECH_COLUMNS = [
    "__case_id",
    "__source",
    "__error_type",
    "__entity_type_detected",
    "__entity_type_expected",
    "__status",
    "__policy_action",
    "__safe_context",
    "__real_value_present",
    "__sensitivity",
    "__recommended_decision",
    "__review_reason",
]


ALL_COLUMNS = VISIBLE_COLUMNS + TECH_COLUMNS


DECISION_LABEL_TO_CODE = {
    "Принять рекомендацию системы": "accept",
    "Скрыть один раз": "hide_once",
    "Скрывать как подозрительное": "hide_as_suspect",
    "Добавить в manual_hide / скрывать": "add_to_manual_hide",
    "Добавить в allow / не скрывать": "add_to_manual_allow",
    "Публичный орган / оставить": "mark_public_org",
    "Частная организация / скрывать": "mark_private_org",
    "Создать правило-кандидат": "create_candidate_rule",
    "Отклонить кейс": "reject_case",
    "Объединить алиасы": "merge_aliases",
}

DECISION_CODE_TO_LABEL = {v: k for k, v in DECISION_LABEL_TO_CODE.items()}


SCOPE_LABEL_TO_CODE = {
    "Один раз": "one_time",
    "В этом проекте": "project_local",
    "Кандидат в общее правило": "global_candidate",
    "Активное общее правило": "global_active",
    "Отклонено": "rejected",
}

SCOPE_CODE_TO_LABEL = {v: k for k, v in SCOPE_LABEL_TO_CODE.items()}


YES_NO_LABEL_TO_BOOL = {
    "Да": True,
    "Нет": False,
}

BOOL_TO_YES_NO_LABEL = {
    True: "Да",
    False: "Нет",
}


TARGET_TYPES = [
    "INN",
    "OCR_SUSPECT_INN",
    "PASSPORT",
    "PHONE",
    "ADDRESS_DETAIL",
    "PERSON",
    "ORG_PRIVATE",
    "PUBLIC_ORG",
    "SERVICE_NUMBER",
    "UNKNOWN",
]


SAMPLE_CASES = [
    {
        "case_id": "case_syn_0001",
        "source": "synthetic",
        "file_name": "synthetic_inn_cases",
        "block": "case_001",
        "error_type": "VALID_ENTITY",
        "entity_type_detected": "INN",
        "entity_type_expected": "INN",
        "status": "VALID_ENTITY",
        "policy_action": "MASK_CONFIDENTLY",
        "safe_context": "ИНН: 473254765214",
        "real_value_present": False,
        "sensitivity": "synthetic",
        "review_reason": "valid_inn_length",
        "recommended_decision": "accept",
        "created_at": datetime.now().isoformat(timespec="seconds"),
    },
    {
        "case_id": "case_syn_0002",
        "source": "synthetic",
        "file_name": "synthetic_inn_cases",
        "block": "case_002",
        "error_type": "OCR_SUSPECT",
        "entity_type_detected": "",
        "entity_type_expected": "OCR_SUSPECT_INN",
        "status": "SUSPECT_ENTITY",
        "policy_action": "MASK_AND_REVIEW",
        "safe_context": "ИНН: <DIGITS_13>",
        "real_value_present": False,
        "sensitivity": "synthetic",
        "review_reason": "invalid_inn_length_near_inn_label",
        "recommended_decision": "hide_as_suspect",
        "created_at": datetime.now().isoformat(timespec="seconds"),
    },
    {
        "case_id": "case_syn_0003",
        "source": "synthetic",
        "file_name": "synthetic_inn_cases",
        "block": "case_003",
        "error_type": "FALSE_POSITIVE_RISK",
        "entity_type_detected": "SERVICE_NUMBER",
        "entity_type_expected": "SERVICE_NUMBER",
        "status": "SKIPPED_BY_ALLOW",
        "policy_action": "ALLOW",
        "safe_context": "АКТ № <DIGITS_13>",
        "real_value_present": False,
        "sensitivity": "synthetic",
        "review_reason": "service_number_context",
        "recommended_decision": "add_to_manual_allow",
        "created_at": datetime.now().isoformat(timespec="seconds"),
    },
]


def ensure_structure():
    for d in [
        FEEDBACK_DIR,
        REVIEW_DIR,
        EXPECTED_DIR,
        RULES_DIR,
        CORPUS_SYNTHETIC_DIR,
        ROOT / "output" / "reports",
        BACKUP_DIR,
    ]:
        d.mkdir(parents=True, exist_ok=True)

    for p in [
        CASES_PATH,
        DECISIONS_PATH,
        CANDIDATE_RULES_PATH,
        BASELINE_REGRESSION_CASES_PATH,
        REVIEW_GENERATED_CASES_PATH,
        MANUAL_HIDE_PATH,
        MANUAL_ALLOW_PATH,
    ]:
        if not p.exists():
            p.write_text("", encoding="utf-8")


def read_jsonl(path: Path):
    if not path.exists():
        return []

    rows = []
    with path.open("r", encoding="utf-8") as f:
        for line_no, line in enumerate(f, start=1):
            line = line.strip()
            if not line:
                continue

            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError as e:
                print(f"Ошибка JSON в {path}, строка {line_no}: {e}")

    return rows


def write_jsonl(path: Path, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def append_unique_line(path: Path, line: str):
    line = (line or "").strip()
    if not line:
        return

    existing = set()
    if path.exists():
        existing = {
            x.strip()
            for x in path.read_text(encoding="utf-8").splitlines()
            if x.strip()
        }

    if line not in existing:
        with path.open("a", encoding="utf-8") as f:
            f.write(line + "\n")


def backup_file(path: Path):
    if not path.exists() or path.stat().st_size == 0:
        return None

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    backup_path = BACKUP_DIR / f"{path.stem}_backup_{timestamp}{path.suffix}"
    shutil.copy2(path, backup_path)
    return backup_path


def seed_cases():
    ensure_structure()

    existing = read_jsonl(CASES_PATH)
    if existing:
        print("feedback/cases.jsonl уже содержит кейсы. Тестовые кейсы не добавляю.")
        return

    write_jsonl(CASES_PATH, SAMPLE_CASES)
    print("Добавлены стартовые синтетические кейсы в feedback/cases.jsonl")


def add_validation(ws, column_name, values, start_row=2, end_row=1000):
    col_index = ALL_COLUMNS.index(column_name) + 1
    col_letter = get_column_letter(col_index)

    formula = '"' + ",".join(values) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)

    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def autosize(ws):
    for col_idx, column_name in enumerate(ALL_COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(column_name)

        for cell in ws[col_letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, min(len(str(cell.value)), 80))

        if column_name.startswith("__"):
            ws.column_dimensions[col_letter].hidden = True
            ws.column_dimensions[col_letter].width = 3
        else:
            ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 60))


def suggested_label(case):
    code = case.get("recommended_decision") or ""

    if code in DECISION_CODE_TO_LABEL:
        return DECISION_CODE_TO_LABEL[code]

    status = case.get("status")
    action = case.get("policy_action")

    if status == "SUSPECT_ENTITY" or action == "MASK_AND_REVIEW":
        return DECISION_CODE_TO_LABEL["hide_as_suspect"]

    if action == "ALLOW" or status == "SKIPPED_BY_ALLOW":
        return DECISION_CODE_TO_LABEL["add_to_manual_allow"]

    return DECISION_CODE_TO_LABEL["accept"]


def suggested_scope_label(case):
    status = case.get("status")
    if status == "SUSPECT_ENTITY":
        return SCOPE_CODE_TO_LABEL["global_candidate"]
    return SCOPE_CODE_TO_LABEL["project_local"]


def export_review_table():
    ensure_structure()

    cases = read_jsonl(CASES_PATH)
    if not cases:
        print("Нет кейсов в feedback/cases.jsonl.")
        print("Сначала выполни: python review_tool.py setup")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Решения"

    ws.append(ALL_COLUMNS)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, case in enumerate(cases, start=1):
        type_value = case.get("entity_type_expected") or case.get("entity_type_detected") or ""
        system_suggestion = (
            f"{case.get('status', '')} / {case.get('policy_action', '')} / {type_value}"
        ).strip(" /")

        visible = {
            "№": idx,
            "Кейс": case.get("case_id", ""),
            "Файл": case.get("file_name", ""),
            "Блок": case.get("block", ""),
            "Фрагмент для проверки": case.get("safe_context", ""),
            "Что предлагает система": system_suggestion,
            "Почему попало в проверку": case.get("review_reason", ""),
            "Решение человека": suggested_label(case),
            "Область действия": suggested_scope_label(case),
            "Тип результата": type_value or "UNKNOWN",
            "Создать правило-кандидат?": "Да" if case.get("status") in ["SUSPECT_ENTITY", "NEEDS_REVIEW"] else "Нет",
            "Создать review-test?": "Да",
            "Комментарий человека": "",
        }

        tech = {
            "__case_id": case.get("case_id", ""),
            "__source": case.get("source", ""),
            "__error_type": case.get("error_type", ""),
            "__entity_type_detected": case.get("entity_type_detected", ""),
            "__entity_type_expected": case.get("entity_type_expected", ""),
            "__status": case.get("status", ""),
            "__policy_action": case.get("policy_action", ""),
            "__safe_context": case.get("safe_context", ""),
            "__real_value_present": case.get("real_value_present", ""),
            "__sensitivity": case.get("sensitivity", ""),
            "__recommended_decision": case.get("recommended_decision", ""),
            "__review_reason": case.get("review_reason", ""),
        }

        row = []
        for col in ALL_COLUMNS:
            value = visible.get(col, tech.get(col, ""))
            if isinstance(value, (list, dict)):
                value = json.dumps(value, ensure_ascii=False)
            row.append(value)

        ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_name in [
        "Фрагмент для проверки",
        "Что предлагает система",
        "Почему попало в проверку",
        "Комментарий человека",
    ]:
        col_idx = ALL_COLUMNS.index(col_name) + 1
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    add_validation(ws, "Решение человека", list(DECISION_LABEL_TO_CODE.keys()))
    add_validation(ws, "Область действия", list(SCOPE_LABEL_TO_CODE.keys()))
    add_validation(ws, "Тип результата", TARGET_TYPES)
    add_validation(ws, "Создать правило-кандидат?", ["Да", "Нет"])
    add_validation(ws, "Создать review-test?", ["Да", "Нет"])

    autosize(ws)

    help_ws = wb.create_sheet("README")
    help_ws["A1"] = "Как работать с таблицей"
    help_ws["A1"].font = Font(bold=True, size=14)

    help_lines = [
        "1. Открой лист 'Решения'.",
        "2. Проверь фрагмент, статус и причину review.",
        "3. В колонке 'Решение человека' выбери действие.",
        "4. В колонке 'Область действия' выбери scope.",
        "5. 'Создать правило-кандидат?' = Да, если кейс может стать правилом.",
        "6. 'Создать review-test?' = Да, если нужно сохранить проверку из этого кейса.",
        "7. Сохрани файл и закрой Excel.",
        "8. Выполни: python review_tool.py import",
        "",
        "Важно: import пишет review-tests в expected/review_generated_cases.jsonl.",
        "Основной baseline expected/regression_cases.jsonl не перезаписывается.",
        "",
        "Если таблица содержит реальные значения, файл считается sensitive.",
    ]

    for i, line in enumerate(help_lines, start=3):
        help_ws[f"A{i}"] = line

    wb.save(REVIEW_PATH)
    print(f"Создана таблица: {REVIEW_PATH}")


def bool_from_label(value):
    if value is None:
        return False
    return YES_NO_LABEL_TO_BOOL.get(str(value).strip(), False)


def load_review_rows():
    if not REVIEW_PATH.exists():
        raise FileNotFoundError(
            f"Не найден файл {REVIEW_PATH}. Сначала выполни: python review_tool.py export"
        )

    wb = load_workbook(REVIEW_PATH)
    if "Решения" not in wb.sheetnames:
        raise ValueError("В Excel не найден лист 'Решения'.")

    ws = wb["Решения"]

    headers = [cell.value for cell in ws[1]]
    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        item = dict(zip(headers, row))
        if not item.get("__case_id") and not item.get("Кейс"):
            continue
        rows.append(item)

    return rows


def build_decision(row):
    decision_label = row.get("Решение человека") or ""
    scope_label = row.get("Область действия") or ""

    decision_code = DECISION_LABEL_TO_CODE.get(decision_label, str(decision_label).strip())
    scope_code = SCOPE_LABEL_TO_CODE.get(scope_label, str(scope_label).strip())

    return {
        "case_id": row.get("__case_id") or row.get("Кейс"),
        "decision": decision_code,
        "scope": scope_code,
        "target_type": row.get("Тип результата") or row.get("__entity_type_expected") or "",
        "create_candidate_rule": bool_from_label(row.get("Создать правило-кандидат?")),
        "create_review_test": bool_from_label(row.get("Создать review-test?")),
        "human_reason": row.get("Комментарий человека") or "",
        "safe_context": row.get("__safe_context") or row.get("Фрагмент для проверки") or "",
        "decided_at": datetime.now().isoformat(timespec="seconds"),
        "created_by": "review_table",
    }


def build_candidate_rule(row, decision):
    case_id = decision["case_id"]
    target_type = decision["target_type"] or "UNKNOWN"
    reason = row.get("__review_reason") or row.get("Почему попало в проверку") or ""
    decision_code = decision["decision"]

    return {
        "rule_id": f"candidate_{case_id}",
        "based_on_case_id": case_id,
        "rule_type": "human_review_candidate",
        "target_type": target_type,
        "description": (
            f"Правило-кандидат на основе {case_id}: "
            f"если найден кейс типа {target_type} с причиной {reason}, "
            f"применить решение {decision_code}."
        ),
        "pseudo_logic": (
            f"if review_reason == '{reason}' "
            f"and target_type == '{target_type}' "
            f"=> decision '{decision_code}'"
        ),
        "status": "candidate",
        "requires_regression": True,
        "risk": "Перед продвижением в active rules нужны positive/negative regression tests.",
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }


def build_review_test(row, decision):
    case_id = decision["case_id"]
    safe_context = decision.get("safe_context") or ""
    target_type = decision.get("target_type") or "UNKNOWN"
    decision_code = decision.get("decision") or ""
    status = row.get("__status") or ""
    policy_action = row.get("__policy_action") or ""

    expected_contains = []
    expected_not_contains = []

    if decision_code in ["hide_as_suspect", "hide_once", "add_to_manual_hide", "accept"]:
        # accept = принять рекомендацию системы. Для MASK_* ожидаем токен.
        if policy_action.startswith("MASK"):
            expected_contains = [f"[{target_type}_1]"]
        else:
            expected_contains = [safe_context]
    elif decision_code in ["add_to_manual_allow", "mark_public_org", "reject_case"]:
        expected_contains = [safe_context]

    return {
        "test_id": f"review_test_{case_id}",
        "based_on_case_id": case_id,
        "input": safe_context,
        "safe_context": safe_context,
        "expected_contains": expected_contains,
        "expected_not_contains": expected_not_contains,
        "expected_status": status,
        "expected_policy_action": policy_action,
        "test_type": "review_generated",
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }


def import_review_decisions():
    ensure_structure()

    rows = load_review_rows()

    decisions = []
    candidate_rules = []
    review_tests = []

    for row in rows:
        decision = build_decision(row)

        if not decision["decision"]:
            continue

        decisions.append(decision)

        if decision["create_candidate_rule"]:
            candidate_rules.append(build_candidate_rule(row, decision))

        if decision["create_review_test"]:
            review_tests.append(build_review_test(row, decision))

        safe_context = decision.get("safe_context") or ""
        target_type = decision.get("target_type") or ""
        decision_code = decision.get("decision") or ""

        if decision_code == "add_to_manual_hide":
            append_unique_line(MANUAL_HIDE_PATH, f"{target_type}: {safe_context}")

        if decision_code in ["add_to_manual_allow", "mark_public_org"]:
            append_unique_line(MANUAL_ALLOW_PATH, safe_context)

    for path in [DECISIONS_PATH, CANDIDATE_RULES_PATH, REVIEW_GENERATED_CASES_PATH]:
        backup = backup_file(path)
        if backup:
            print(f"Backup создан: {backup.relative_to(ROOT)}")

    write_jsonl(DECISIONS_PATH, decisions)
    write_jsonl(CANDIDATE_RULES_PATH, candidate_rules)
    write_jsonl(REVIEW_GENERATED_CASES_PATH, review_tests)

    print(f"Импортировано решений: {len(decisions)}")
    print(f"Создано candidate-rules: {len(candidate_rules)}")
    print(f"Создано review-generated tests: {len(review_tests)}")
    print("")
    print("ВАЖНО: expected/regression_cases.jsonl не изменялся.")
    print(f"Review-tests записаны в: {REVIEW_GENERATED_CASES_PATH.relative_to(ROOT)}")


def count_jsonl(path: Path):
    return len(read_jsonl(path))


def show_status():
    ensure_structure()

    files = [
        CASES_PATH,
        REVIEW_PATH,
        DECISIONS_PATH,
        CANDIDATE_RULES_PATH,
        BASELINE_REGRESSION_CASES_PATH,
        REVIEW_GENERATED_CASES_PATH,
        MANUAL_HIDE_PATH,
        MANUAL_ALLOW_PATH,
    ]

    print("Статус review-контура:")
    for p in files:
        exists = p.exists()
        size = p.stat().st_size if exists else 0
        count = ""
        if p.suffix == ".jsonl" and exists:
            count = f", lines={count_jsonl(p)}"
        print(f"- {p.relative_to(ROOT)}: {'есть' if exists else 'нет'}, {size} байт{count}")


def main():
    import sys

    print("=== REVIEW TOOL v3 / split baseline vs review-generated tests ===")

    command = sys.argv[1].strip().lower() if len(sys.argv) > 1 else "help"

    if command == "setup":
        ensure_structure()
        seed_cases()
        show_status()

    elif command == "export":
        export_review_table()
        show_status()

    elif command == "import":
        import_review_decisions()
        show_status()

    elif command == "status":
        show_status()

    else:
        print("Команды:")
        print("  python review_tool.py setup")
        print("  python review_tool.py export")
        print("  python review_tool.py import")
        print("  python review_tool.py status")
        print("")
        print("Важно:")
        print("  expected/regression_cases.jsonl — основной synthetic baseline")
        print("  expected/review_generated_cases.jsonl — тесты из Excel-review")


if __name__ == "__main__":
    main()
